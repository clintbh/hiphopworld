import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
import pandas as pd

# Load the workbook
workbook = load_workbook(filename='list.xlsx')

# Get the worksheet
worksheet = workbook['Sheet1']

# Get the terms from column A of the worksheet
terms = [row[0] for row in worksheet.iter_rows(min_row=2, min_col=1, values_only=True)]
print(f"Number of terms: {len(terms)}")


# Create a new workbook
workbook = Workbook()

# Get the active worksheet
worksheet = workbook.active

# Set the column headers
worksheet.cell(row=1, column=1, value='Term')
worksheet.cell(row=1, column=2, value='Year started')
worksheet.cell(row=1, column=3, value='Location of birth')
worksheet.cell(row=1, column=4, value='Origin')

# Loop through the terms and search Wikipedia for each one
for i, term in enumerate(terms, start=2):
    # Construct the URL for the Wikipedia page
    url = f'https://en.wikipedia.org/wiki/{term}'

    try:
        # Send a GET request to the URL and store the response
        response = requests.get(url)

        # Create a BeautifulSoup object from the response text
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the infobox on the page (either class name will work)
        infobox = soup.find('table', {'class': 'infobox biography vcard'}) or \
                  soup.find('table', {'class': 'infobox vcard plainlist'})

        # Extract the year started and location of birth from the infobox (if it exists)
        year_started = ''
        location_of_birth = ''
        origin = None
        if infobox is not None:
            for row in infobox.find_all('tr'):
                if row.th and 'Origin' in row.th.text:
                    origin = row.td.text.strip()
                if row.th and 'Years' in row.th.text:
                    year_started = row.td.text.strip()
                if row.th and 'Born' in row.th.text:
                    location_of_birth = row.td.text.strip()

        # If the year_started and location_of_birth results are null, try adding "_rapper" to the term
        if not year_started and not location_of_birth:
            term = f'{term}_(rapper)'
            url = f'https://en.wikipedia.org/wiki/{term}'
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            infobox = soup.find('table', {'class': 'infobox biography vcard'}) or \
                      soup.find('table', {'class': 'infobox vcard plainlist'})
            if infobox is not None:
                for row in infobox.find_all('tr'):
                    if row.th and 'Origin' in row.th.text:
                        origin = row.td.text.strip()
                    if row.th and 'Years' in row.th.text:
                        year_started = row.td.text.strip()
                    if row.th and 'Born' in row.th.text:
                        location_of_birth = row.td.text.strip()

        # Write the results to the worksheet
        worksheet.cell(row=i, column=1, value=term)
        worksheet.cell(row=i, column=2, value=year_started)
        worksheet.cell(row=i, column=3, value=location_of_birth)
        if origin is not None:
            worksheet.cell(row=i, column=4, value=origin)
        else:
            worksheet.cell(row=i, column=4, value='')
    except Exception as e:
        print(f"Error: {e}. Skipping {term}")
        continue

# Save the workbook
workbook.save('origin3.xlsx')
